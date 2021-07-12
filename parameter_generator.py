import random
import openpyxl
parameters=openpyxl.Workbook()
parameters.create_sheet('parameters_random')
prm=parameters['parameters_random']
for i in range(1,2501):
    prm.cell(row=i, column=1).value = random.uniform(350, 700)
    prm.cell(row=i, column=2).value = random.uniform(0.5, 5)
    prm.cell(row=i, column=3).value = random.uniform(0.5/197, 2/197)
    #prm.cell(row=i, column=3).value = random.uniform(0.5 / 1000,2 / 1000)
    prm.cell(row=i, column=4).value = random.uniform(0.1, 1)
    prm.cell(row=i, column=5).value = random.uniform(0.5/197, 2/197)
    #prm.cell(row=i, column=5).value = random.uniform(0.5 / 1000, 2 / 1000)
parameters.save('parameters_random.xlsx')