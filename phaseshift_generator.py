import numpy as np
import math
from scipy import integrate
import openpyxl

m = 138.5
def phaseshift_g(E,m0,g,c,G,d):
    k_on = (E**2/4-m**2)**0.5
    rou = math.pi*k_on*E/4
    T = -rou*t(k_on,k_on,E,m0,g,c,G,d)
    S = 1+(2j*T)
    ps = np.angle(S)/2
    return ps



def g1(k,c,g):
    return g/m**0.5 * 1/(1+(c*k)**2)



def g2(k,d):
    return 1/(1+(d*k)**2)**2



def t(k1,k2,E,m0,g,c,G,d):
    k_on = (E**2/4-m**2)**0.5
    rou = math.pi*k_on*E/4
    f11 = lambda x: g1(x,c,g) * g1(x,c,g) * x ** 2 / (E - 2 * (m ** 2 + x ** 2) ** 0.5 ) + g1(k_on,c,g) * g1(k_on,c,g) * k_on ** 2 / (x ** 2 - k_on ** 2) * E / 2
    g11 = integrate.quad(f11, 0, 1e6)[0].real - 1j*g1(k_on,c,g)*g1(k_on,c,g)*rou
    f12 = lambda x: g1(x,c,g) * g2(x,d) * x ** 2 / (E - 2 * (m ** 2 + x ** 2) ** 0.5 ) + g1(k_on,c,g) * g2(k_on,d) * k_on ** 2 / (x ** 2 - k_on ** 2) * E / 2
    g12 = integrate.quad(f12, 0, 1e6)[0].real - 1j*g1(k_on,c,g)*g2(k_on,d)*rou
    f22 = lambda x: g2(x,d) * g2(x,d) * x ** 2 / (E - 2 * (m ** 2 + x ** 2) ** 0.5 ) + g2(k_on,d) * g2(k_on,d) * k_on ** 2 / (x ** 2 - k_on ** 2) * E / 2
    g22 = integrate.quad(f22, 0, 1e6)[0].real - 1j*g2(k_on,d)*g2(k_on,d)*rou
    N = -g12**2 + (E-g11-m0) * (-g22+m**2/G)
    t11 = (m**2/G - g22)/N
    t12 = g12 / N
    t21 = g12 / N
    t22 = (E - g11 - m0) / N
    return g1(k1,c,g)*t11*g1(k2,c,g) + g1(k1,c,g)*t12*g2(k2,d) + g2(k1,d)*t21*g1(k2,c,g) + g2(k1,d)*t22*g2(k2,d)





phaseshift = openpyxl.Workbook()
phaseshift.create_sheet('phaseshift')
pha=phaseshift['phaseshift']

parameters = openpyxl.load_workbook('parameters_random.xlsx')
para = parameters['parameters_random']
for i in range(195,196):
    m0 = para.cell(row = i, column = 1).value
    g = para.cell(row = i, column = 2).value
    c = para.cell(row = i, column = 3).value
    G = para.cell(row = i, column = 4).value
    d = para.cell(row = i, column = 5).value
    j = 0
    zeros = phaseshift_g(277+1e-4,m0,g,c,G,d)
    term0 = zeros
    for E in np.linspace(277+1e-4,1000,42):
        j += 1
        term1 = phaseshift_g(E,m0,g,c,G,d)
        if term0 - term1 >= 20  and ((term1 > 160 and term1 < 200) or (term1 > -200 and term1 < -160) or (term1 > -20 and term1 < 20)):
            pha.cell(row=i, column=j).value = term1 + 180 - zeros
        elif term1 - term0 >= 20 and ((term1 > 160 and term1 < 200) or (term1 > -200 and term1 < -160) or (term1 > -20 and term1 < 20)):
            pha.cell(row=i, column=j).value = term1 - 180 - zeros
        else:
            pha.cell(row=i, column=j).value = term1 - zeros
        term0 = pha.cell(row=i, column=j).value + zeros
phaseshift.save('phaseshift.xlsx')