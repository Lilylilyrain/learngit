import openpyxl
from scipy.optimize import root
from difference import five_point_formula
import numpy as np
import math
import random
from scipy import integrate
import matplotlib.pyplot as plt

m = 138.5
N = 197


def g1(k, c, g):
    return g / m ** 0.5 * 1 / (1 + (c * k) ** 2)


def g2(k, d):
    return 1 / (1 + (d * k) ** 2) ** 2


def trans(x):
    return (x + 1) / (1 - x) + 1e9



def f11(E, g, c, k_on):
    def f(x):
        return g1(x, c, g) * g1(x, c, g) * x ** 2 / (E - 2 * (m ** 2 + x ** 2) ** 0.5)

    return f



def f12(E, g, c, d, k_on):
    def f(x):
        return g1(x, c, g) * g2(x, d) * x ** 2 / (E - 2 * (m ** 2 + x ** 2) ** 0.5)

    return f


def f22(E, d, k_on):
    def f(x):
        return g2(x, d) * g2(x, d) * x ** 2 / (E - 2 * (m ** 2 + x ** 2) ** 0.5)

    return f


def quad(f, x1, y1, x2, y2):
    k = (y2 - y1) / (x2 - x1)
    f11_r = lambda x: f(x + (k * (x - x1) + y1) * 1j).real
    f11_i = lambda x: f(x + (k * (x - x1) + y1) * 1j).imag
    result = integrate.quad(f11_r, x1, x2)[0] + 1j * integrate.quad(f11_i, x1, x2)[0]
    return result


def det(E, m0, g, c, G, d):
    k_on = (E ** 2 / 4 - m ** 2) ** 0.5
    theta = np.angle(k_on) - math.pi / 13
    # print(theta)
    # theta = math.pi/4
    rou = math.pi * k_on * E / 4
    # g11 = gauss(f11(E,g,c,k_on),theta)
    # g12 = gauss(f12(E,g,c,d,k_on),theta)
    # g22 = gauss(f22(E,d,k_on),theta)
    # print("gauss",g22)
    epsilon = 1e-5
    f11_r = lambda x: (g1(x * math.cos(theta) + 1j * x * math.sin(theta), c, g) * g1(
        x * math.cos(theta) + 1j * x * math.sin(theta), c, g) * (
                                   x * math.cos(theta) + 1j * x * math.sin(theta)) ** 2 / (E - 2 * (
                m ** 2 + (x * math.cos(theta) + 1j * x * math.sin(theta)) ** 2) ** 0.5 + 1j * epsilon)).real
    f11_i = lambda x: (g1(x * math.cos(theta) + 1j * x * math.sin(theta), c, g) * g1(
        x * math.cos(theta) + 1j * x * math.sin(theta), c, g) * (
                                   x * math.cos(theta) + 1j * x * math.sin(theta)) ** 2 / (E - 2 * (
                m ** 2 + (x * math.cos(theta) + 1j * x * math.sin(theta)) ** 2) ** 0.5 + 1j * epsilon)).imag
    g11 = integrate.quad(f11_r, 0, np.inf)[0] + 1j * integrate.quad(f11_i, 0, np.inf)[0]
    f12_r = lambda x: (g1(x * math.cos(theta) + 1j * x * math.sin(theta), c, g) * g2(
        x * math.cos(theta) + 1j * x * math.sin(theta), d) * (x * math.cos(theta) + 1j * x * math.sin(theta)) ** 2 / (
                                   E - 2 * (m ** 2 + (x * math.cos(theta) + 1j * x * math.sin(
                               theta)) ** 2) ** 0.5 + 1j * epsilon)).real
    f12_i = lambda x: (g1(x * math.cos(theta) + 1j * x * math.sin(theta), c, g) * g2(
        x * math.cos(theta) + 1j * x * math.sin(theta), d) * (x * math.cos(theta) + 1j * x * math.sin(theta)) ** 2 / (
                                   E - 2 * (m ** 2 + (x * math.cos(theta) + 1j * x * math.sin(
                               theta)) ** 2) ** 0.5 + 1j * epsilon)).imag
    g12 = integrate.quad(f12_r, 0, np.inf)[0] + 1j * integrate.quad(f12_i, 0, np.inf)[0]
    f22_r = lambda x: (g2(x * math.cos(theta) + 1j * x * math.sin(theta), d) * g2(
        x * math.cos(theta) + 1j * x * math.sin(theta), d) * (x * math.cos(theta) + 1j * x * math.sin(theta)) ** 2 / (
                                   E - 2 * (m ** 2 + (x * math.cos(theta) + 1j * x * math.sin(
                               theta)) ** 2) ** 0.5 + 1j * epsilon)).real
    f22_i = lambda x: (g2(x * math.cos(theta) + 1j * x * math.sin(theta), d) * g2(
        x * math.cos(theta) + 1j * x * math.sin(theta), d) * (x * math.cos(theta) + 1j * x * math.sin(theta)) ** 2 / (
                                   E - 2 * (m ** 2 + (x * math.cos(theta) + 1j * x * math.sin(
                               theta)) ** 2) ** 0.5 + 1j * epsilon)).imag
    g22 = integrate.quad(f22_r, 0, np.inf)[0] + 1j * integrate.quad(f22_i, 0, np.inf)[0]
    # print("scipy:", g22)
    # result = 1-G*g22-g11/(E-m0)-G*g12**2/(E-m0)+G*g11*g22/(E-m0)
    result = E - m0 - G * g22 - g11 - G * g12 ** 2 + G * g11 * g22
    return result


def root_r(m0, g, c, G, d):
    def root(E):
        return det(E, m0, g, c, G, d).real

    return root


def root_i(m0, g, c, G, d):
    def root(E):
        return det(E, m0, g, c, G, d).imag

    return root


def root_gradient_r_0(m0, g, c, G, d):
    def gradient(E):
        h = 1e-6
        # return (f(x - 2 * h) - 8 * f(x - h) + 8 * f(x + h) - f(x + 2 * h)) / 12 / h
        return (root_r(m0, g, c, G, d)(E[0] - 2 * h + E[1] * 1j) - 8 * root_r(m0, g, c, G, d)(
            E[0] - h + E[1] * 1j) + 8 * root_r(m0, g, c, G, d)(E[0] + h + E[1] * 1j) - root_r(m0, g, c, G, d)(
            E[0] + 2 * h + E[1] * 1j)) / 12 / h

    return gradient


def root_gradient_r_1(m0, g, c, G, d):
    def gradient(E):
        h = 1e-6
        return (root_r(m0, g, c, G, d)(E[0] + (E[1] - 2 * h) * 1j) - 8 * root_r(m0, g, c, G, d)(
            E[0] + (E[1] - h) * 1j) + 8 * root_r(m0, g, c, G, d)(E[0] + (E[1] + h) * 1j) - root_r(m0, g, c, G, d)(
            E[0] + (E[1] + 2 * h) * 1j)) / 12 / h

    return gradient


def root_gradient_i_0(m0, g, c, G, d):
    def gradient(E):
        h = 1e-6
        # return (f(x - 2 * h) - 8 * f(x - h) + 8 * f(x + h) - f(x + 2 * h)) / 12 / h
        return (root_i(m0, g, c, G, d)(E[0] - 2 * h + E[1] * 1j) - 8 * root_i(m0, g, c, G, d)(
            E[0] - h + E[1] * 1j) + 8 * root_i(m0, g, c, G, d)(E[0] + h + E[1] * 1j) - root_i(m0, g, c, G, d)(
            E[0] + 2 * h + E[1] * 1j)) / 12 / h

    return gradient


def root_gradient_i_1(m0, g, c, G, d):
    def gradient(E):
        h = 1e-6
        # return (f(x - 2 * h) - 8 * f(x - h) + 8 * f(x + h) - f(x + 2 * h)) / 12 / h
        return (root_i(m0, g, c, G, d)(E[0] + (E[1] - 2 * h) * 1j) - 8 * root_i(m0, g, c, G, d)(
            E[0] + (E[1] - h) * 1j) + 8 * root_i(m0, g, c, G, d)(E[0] + (E[1] + h) * 1j) - root_i(m0, g, c, G, d)(
            E[0] + (E[1] + 2 * h) * 1j)) / 12 / h

    return gradient


def you_need_my_root(m0, g, c, G, d):
    def func(x):
        E = x[0] + 1j * x[1]
        f = np.array([root_r(m0, g, c, G, d)(E), root_i(m0, g, c, G, d)(E)])
        df = np.array([[root_gradient_r_0(m0, g, c, G, d)(x), root_gradient_r_1(m0, g, c, G, d)(x)],
                       [root_gradient_i_0(m0, g, c, G, d)(x), root_gradient_i_1(m0, g, c, G, d)(x)]])
        return f, df

    return func


def ppp(x):
    return x[0] + x[1]


def pole_position_g1(m0, g, c, G, d):
    sol = root(you_need_my_root(m0, g, c, G, d), np.array([300, -100]), jac=True, method='lm')
    # sol = root(you_need_my_root(m0,g,c,G,d),np.array([300,-100]),jac=True,method='hybr')
    # print(sol)
    return sol.x[0] + sol.x[1] * 1j


def pole_position_g2(m0, g, c, G, d):
    tol = 1e-6
    delta = 0.01
    step = 1
    # initial_r = random.uniform(0,1000)
    # initial_i = random.uniform(-500,0)
    initial_r = 421
    initial_i = -80
    E = initial_r + 1j * initial_i
    limit = 2e3
    E_r = initial_r
    E_i = initial_i
    while abs(det(E, m0, g, c, G, d)) > tol:
        E_old = np.array((E_r, E_i)).T
        x = E_old.T
        f = np.array((root_r(m0, g, c, G, d)(E), root_i(m0, g, c, G, d)(E))).T
        df = np.array(((root_gradient_r_0(m0, g, c, G, d)(x), root_gradient_r_1(m0, g, c, G, d)(x)),
                       (root_gradient_i_0(m0, g, c, G, d)(x), root_gradient_i_1(m0, g, c, G, d)(x))))
        xrr = root_gradient_r_0(m0, g, c, G, d)(x)
        xri = root_gradient_r_1(m0, g, c, G, d)(x)

        xir = root_gradient_i_0(m0, g, c, G, d)(x)
        xii = root_gradient_i_1(m0, g, c, G, d)(x)
        xxx = xrr * xii - xri * xir

        # print(df,np.linalg.det(df))
        if np.linalg.det(df) < 1e-14:
            E_r = random.uniform(500, 1000)
            E_i = random.uniform(-500, 0)
            E = E_r + 1j * E_i
            # print("111",E,det(E,m0,g,c,G,d))
        else:
            print(E_old, xxx, "aaa")
            df_inv = np.linalg.inv(df)
            E_new = E_old - df_inv @ f
            E_r = E_new.T[0]
            E_i = E_new.T[1]
            E_r = E_old.T[0] - (xii * E_old.T[0] - xri * E_old.T[1]) / xxx
            E_i = E_old.T[1] - (-xir * E_old.T[0] + xrr * E_old.T[1]) / xxx
            E = E_r + 1j * E_i
            print(E, det(E, m0, g, c, G, d), xxx, "aaa")
            if abs(E_r) > limit or abs(E_i) > limit:
                E_r = random.uniform(500, 1000)
                E_i = random.uniform(-500, 0)
                E = E_r + 1j * E_i
                # print("222", E, det(E, m0, g, c, G, d))
    return E




def check(a, m0, g, c, G, d):
    resultturple = []
    points = 1500
    for i in range(5):
        r = (i + 1) / 25
        result = 0
        for j in range(points):
            theta = 2 * j * math.pi / points
            x_r = a.real + r * math.cos(theta)
            x_i = a.imag + r * math.sin(theta)
            result += 1 / det(x_r + 1j * x_i, m0, g, c, G, d) * math.pi * 2 * r / points * (
                        math.cos(theta) + 1j * math.sin(theta)) * 1j
            # result += 1/(x_r+1j*x_i-a)*math.pi*2*r/points*(math.cos(theta)+1j*math.sin(theta))*1j
        resultturple.append(result)
    return resultturple






pole_position = openpyxl.Workbook()
pole_position.create_sheet('pole_position')
popo=pole_position['pole_position']
parameters = openpyxl.load_workbook('parameters_random.xlsx')
para = parameters['parameters_random']
for i in range(1,3):
    m0 = para.cell(row = i, column = 1).value
    g = para.cell(row = i, column = 2).value
    c = para.cell(row = i, column = 3).value
    G = para.cell(row = i, column = 4).value
    d = para.cell(row = i, column = 5).value
    popo.cell(row=i, column=1).value = pole_position_g(m0,g,c,G,d).real
    popo.cell(row=i, column=2).value = pole_position_g(m0, g, c, G, d).imag
    print(pole_position_g(m0,g,c,G,d).real)
pole_position.save('pole_position.xlsx')
"""
a_on =  (a**2/4-m**2)**0.5
b = check(a,m0,g,c,G,d)
print(b)
c = det(a,m0,g,c,G,d)
print(c)
"""

"""
if np.linalg.norm(b)>=0.1:
    print("error",i)
if np.linalg.norm(c) >= 0.001:
    print("error",i)
"""


