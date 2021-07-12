import numpy as np
import math
import openpyxl
import scipy.linalg
N = 197
m = 138.5
table = np.zeros(shape=7300)
for i in range(50):
    for j in range(50):
        for k in range(50):
            if i == 0 and j == 0 and k == 0:
                table[i ** 2 + j ** 2 + k ** 2] += 1
            elif (i == 0 and j == 0 and k > 0) or (i == 0 and j > 0 and k == 0) or (i > 0 and j == 0 and k == 0):
                table[i ** 2 + j ** 2 + k ** 2] += 2
            elif (i > 0 and j > 0 and k == 0) or (i > 0 and j == 0 and k > 0) or (i == 0 and j > 0 and k > 0):
                table[i ** 2 + j ** 2 + k ** 2] += 4
            elif i > 0 and j > 0 and k > 0:
                table[i ** 2 + j ** 2 + k ** 2] += 8
#思路：broadcast/元素to列，并行/重构循环\scipy更好，可排序 setsetbyvalue##
def energy_g(l,m0,g,c,G,d):
    l = l / N
    ones = np.ones(shape=(2400, 2400))
    i = np.arange(0, 2400, 1)
    j = np.arange(0, 2400, 1)  # 把行列的指标变成行向量
    ki = (abs(i - 1)) ** 0.5 * 2 * math.pi / l
    ki[0] = 0
    Ki = (ki * ones).T  # 通过广播把ki扩张为矩阵
    first_row = (g / m ** 0.5) * (1 / (1 + (c * ki) ** 2)) * (table[i - 1] / (4 * math.pi)) ** 0.5 * (
                2 * math.pi / l) ** 1.5
    kj = (abs(j - 1)) ** 0.5 * 2 * math.pi / l
    kj[0] = 0
    Kj = kj * ones
    H = (((G / (m ** 2)) * (1 / ((1 + (d * Ki) ** 2) ** 2)) * (1 / ((1 + (d * Kj) ** 2) ** 2)) * (
                table[j - 1] / (4 * math.pi)) ** 0.5 * (2 * math.pi / l) ** 3).T * (
                     table[i - 1] / (4 * math.pi)) ** 0.5).T  # Ki和Kj的矩阵元分别相乘，而Ki中每一行的元素全同，Kj中每一列全同，保证二者各表示行、列指标
    F = np.diag(2 * ((m ** 2 + ki ** 2) ** 0.5))  # F表示自由能对应的对角阵
    for i in range(2400):
       if table[i-1] == 0:
            F[i][i] = 0
    H += F
    H[0] = first_row
    H[:, 0] = first_row.T
    H[0][0] = m0  # 把bare state相关的项单独加上
    results = scipy.linalg.eigvalsh(H,subset_by_value=[276,np.inf])[0:11]
    #for i in range(10):
       # print(H[i][0:10])

    #print(H[4][0:5])
    #for i in range(620):
       # spectrum[i] = abs(spectrum[i])
    #spectrum.sort()
    return results




energy = openpyxl.Workbook()
energy.create_sheet('energy')
eg=energy['energy']
'''
parameters = openpyxl.load_workbook('parameters_random.xlsx')
para = parameters['parameters_random']
for i in range(1114,1115):
    m0 = para.cell(row = i, column = 1).value
    g = para.cell(row = i, column = 2).value
    c = para.cell(row = i, column = 3).value
    G = para.cell(row = i, column = 4).value
    d = para.cell(row = i, column = 5).value
    print(m0,g,c,G,d)
    for l in [11,12,13,14,15,16,17]:
        result = energy_g(l,m0,g,c,G,d)
        for j in range(6):
            #print(l,j,result[j])
            eg.cell(row = i, column = j+1+6*(l-4)).value = result[j]
energy.save('energy_518.xlsx')
'''
result = energy_g(10,404.9062965,3.008403391,0.003627247,0.331229171,0.004090829)
for i in range(7):
    print(result[i])